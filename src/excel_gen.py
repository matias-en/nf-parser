import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Colunas que recebem formato monetário R$
COLUNAS_MOEDA = {
    "Valor Total NF", "BC ICMS", "ICMS", "IPI", "PIS", "COFINS",
    "Valor Frete", "ICMS CTe",
    "Valor Total", "ISSQN", "Valor Líquido",
    "IRRF", "CPP", "CSRF",
    "IRRF Retido", "CPP Retido", "CSLL Retido", "COFINS Retido", "PIS Retido",
    "vProd", "vUnCom", "vTotTrib", "vICMS", "vPIS", "vCOFINS", "vIPI",
    "Total vProd", "Total NF",
}

# Colunas que recebem formato de data
COLUNAS_DATA = {"Data", "Data do Evento", "Mês"}

# Colunas que recebem formato percentual
COLUNAS_PERCENT = {"Aliq ISSQN"}

# Abas com potencial de volume alto — usam write_only via iteração de cursor
ABAS_GRANDES = {"Produtos"}

# Limite de linhas por aba do Excel
LIMITE_LINHAS_EXCEL = 1_048_575 

FORMAT_MOEDA   = 'R$ #,##0.00'
FORMAT_DATA    = 'DD/MM/YYYY'
FORMAT_PERCENT = '0.00%'
FORMAT_INT     = '#,##0'

COR_CABECALHO  = "1F4E79"   # azul escuro
COR_FONTE_CAB  = "FFFFFF"   # branco


def _aplicar_cabecalho(ws, colunas: list):
    """Aplica estilo ao cabeçalho da aba."""
    fill = PatternFill("solid", fgColor=COR_CABECALHO)
    fonte = Font(bold=True, color=COR_FONTE_CAB)
    for col_idx, nome in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=col_idx, value=nome)
        cel.fill = fill
        cel.font = fonte
        cel.alignment = Alignment(horizontal="center")


def _formatar_coluna(ws, col_idx: int, nome_col: str, n_linhas: int):
    if nome_col in COLUNAS_MOEDA:
        fmt = FORMAT_MOEDA
    elif nome_col in COLUNAS_DATA:
        fmt = FORMAT_DATA
    elif nome_col in COLUNAS_PERCENT:
        fmt = FORMAT_PERCENT
    else:
        return 

    for row in range(2, n_linhas + 2):
        ws.cell(row=row, column=col_idx).number_format = fmt


def _ajustar_largura(ws, df: pd.DataFrame):
    for col_idx, col_nome in enumerate(df.columns, start=1):
        try:
            max_conteudo = df[col_nome].astype(str).map(len).max()
        except Exception:
            max_conteudo = 10
        largura = max(max_conteudo, len(str(col_nome))) + 3
        largura = min(largura, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura


def _escrever_aba_normal(wb: Workbook, nome_aba: str, df: pd.DataFrame):
    ws = wb.create_sheet(title=nome_aba)
    colunas = list(df.columns)

    _aplicar_cabecalho(ws, colunas)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, valor in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    n_linhas = len(df)
    for col_idx, nome_col in enumerate(colunas, start=1):
        _formatar_coluna(ws, col_idx, nome_col, n_linhas)

    _ajustar_largura(ws, df)
    logger.info(f"  Aba '{nome_aba}': {n_linhas} linhas gravadas.")


def _escrever_aba_grande(wb: Workbook, nome_aba_base: str, df: pd.DataFrame):
    total = len(df)
    partes = max(1, -(-total // LIMITE_LINHAS_EXCEL))

    for parte in range(partes):
        inicio = parte * LIMITE_LINHAS_EXCEL
        fim = inicio + LIMITE_LINHAS_EXCEL
        df_parte = df.iloc[inicio:fim]

        nome_aba = nome_aba_base if partes == 1 else f"{nome_aba_base} ({parte + 1})"
        _escrever_aba_normal(wb, nome_aba, df_parte)


def salvar_excel(dicionario_dados: dict, nome_arquivo):
    wb = Workbook()
    wb.remove(wb.active)

    escreveu_alguma = False

    for nome_aba, dados in dicionario_dados.items():
        if isinstance(dados, pd.DataFrame):
            df = dados
        elif isinstance(dados, list) and len(dados) > 0:
            df = pd.DataFrame(dados)
        else:
            continue 

        if df.empty:
            continue

        if nome_aba in ABAS_GRANDES:
            _escrever_aba_grande(wb, nome_aba, df)
        else:
            _escrever_aba_normal(wb, nome_aba, df)

        escreveu_alguma = True

    if not escreveu_alguma:
        ws = wb.create_sheet("Aviso")
        ws["A1"] = "Nenhum dado válido encontrado neste lote."
        logger.warning("Nenhum dado válido encontrado — relatório vazio gerado.")

    wb.save(nome_arquivo)
    logger.info(f"Relatório salvo em: {nome_arquivo}")
